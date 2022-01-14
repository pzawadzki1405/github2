# Python program to read an excel file

# import openpyxl module
import openpyxl

file1 = open("MyFile.txt","w")
path = input('Please enter path of excel file')

col_name = 1
col_display = 2
col_bin = 3
col_avaible = 4
col_preffered = 5
col_on_hand = 6
col_UPC = 7

# Give the location of the file
if (path == ''):
    path = "INVENTORY.xlsx"
blacklist = ['96PF250', '96PF444', '96PF251', '96GT15P', '96BOX004', '96FM003A', '96FM003P'
            '96FM005A', '96FM005P', '9696912L', '9696912LB', '9696912M', '9696912MB', '9696912XL',
            '9696912XLB', '9696929L', '9696929LB', '9696929M', '9696929MB', '9696929XL', '9696929XLB',
            '9696929XXL', '9696929XXLB']
# To open the workbook
# workbook object is created
print('Loading the file... Please wait')
wb_obj = openpyxl.load_workbook(path)
print(wb_obj.sheetnames)

def checkifpallet(row_value, column_value):
    if (row_value == 1):
        row_value = row_value + 1 #first is a title
    location_string = str(wb_obj.active.cell(row = row_value, column = column_value).value)
    aisle = int(location_string[0:2])
    if ((aisle >= 1) & (aisle <=9)):
        if (location_string.endswith('E') | location_string.endswith('F')):
            #print('its pallet ' + location_string )
            return True
        else:
            #print('its not pallet ' + location_string )
            return False
    elif ((aisle >= 10) & (aisle <=17)):
        if (location_string.endswith('G') | location_string.endswith('H')):
            #print('its pallet ' + location_string )
            return True
        else:
            #print('its not pallet ' + location_string )
            return False
    elif((aisle >= 18)):
        if (location_string.endswith('F')):
            #print('its pallet ' + location_string )
            return True
        else:
            #print('its not pallet ' + location_string )
            return False
    else:
        return False

    #print(str(aisle))

def testfunc():
    wb_obj.active = 2
    #sheet_obj = wb_obj.active
    for i in range(1, 100):
        checkifpallet(i, 3)
    while(True):
        menu = input('Please enter 0 to go to main menu')
        if (menu == '0'):
            break
        else:
            print('Wrong action, please try again')

def inventory():
    wb_obj.active = 2
    sheet_obj = wb_obj.active
    treshhold_min = input('Please enter minimum quantity on location: ')

    for i in range(1, len(list(sheet_obj.rows))):
        cell_obj = sheet_obj.cell(row = i, column = 1)
        cell_obj_v = cell_obj.value
        cell_obj_next = sheet_obj.cell(row = i+1, column = 1)
        cell_obj_next_v = cell_obj_next.value
        if (cell_obj_v == cell_obj_next_v):
            if (int((sheet_obj.cell(row = i, column = 4)).value) <= int(treshhold_min) ):
                print(str(cell_obj_v) +
                    ' at location ' + str(((sheet_obj.cell(row = i, column = 3)).value)) +
                    ' have ' + str(((sheet_obj.cell(row = i, column = 4)).value)) + ' avaible')
    while(True):
        menu = input('Please enter 0 to go to main menu')
        if (menu == '0'):
            break
        else:
            print('Wrong action, please try again')

def replenishment():
    wb_obj.active = 2
    sheet_obj = wb_obj.active
    treshhold_min = input('Please enter minimum quantity on location: ')
    threshold_max = input('Please enter maximum quantity on location: ')
    file1.write('Minimum quanttity to do replenishment: ' + str(treshhold_min) + '\n')
    file1.write('Maximum quanttity after replenishment: ' + str(threshold_max) + '\n')
    black = False
    for i in range(1, len(list(sheet_obj.rows))):
        cell_obj = sheet_obj.cell(row = i, column = 1)
        cell_obj_v = cell_obj.value
        cell_obj_next = sheet_obj.cell(row = i+1, column = 1)
        cell_obj_next_v = cell_obj_next.value
        if (cell_obj_v == cell_obj_next_v):
            if (int((sheet_obj.cell(row = i, column = 4)).value) <= int(treshhold_min) ):
                result = int((sheet_obj.cell(row = i, column = 4)).value) + int((sheet_obj.cell(row = i+1, column = 4)).value)
                if (result <= int(threshold_max)):
                    for j in range(0, len(blacklist)):
                        if (blacklist[j] == str(cell_obj_v)):
                            black = True
                    if (black == False):
                        message = ('Transfer ' + str(cell_obj_v) +
                            ' from ' + str(((sheet_obj.cell(row = i, column = 3)).value)) +
                            ' to ' + str(((sheet_obj.cell(row = i+1, column = 3)).value)) +
                            ' it will be ' + str(result)  + ' total')
                        print(message)
                        #file1.write((message + '\n'))
                        black = False

    while(True):
        menu = input('Please enter 0 to go to main menu')
        if (menu == '0'):
            file1.close()
            break
        else:
            print('Wrong action, please try again')

def doublelocations():

    wb_obj.active = 2
    sheet_obj = wb_obj.active
    black = False
    save_file = open("double_locations.txt","w")
    for i in range(1, len(list(sheet_obj.rows))):
        cell_obj = sheet_obj.cell(row = i, column = 1)
        cell_obj_v = cell_obj.value
        cell_obj_next = sheet_obj.cell(row = i+1, column = 1)
        cell_obj_next_v = cell_obj_next.value
        if (i>1):
            cell_obj_previous = sheet_obj.cell(row = i-1, column = 1)
            cell_obj_previous_v = cell_obj_previous.value
        else:
            cell_obj_previous = 0
            cell_obj_previous_v = 0
        if ((str(cell_obj_v) == str(cell_obj_next_v)) | (str(cell_obj_v) == str(cell_obj_previous_v))):
            for j in range(0, len(blacklist)):
                if (blacklist[j] == str(cell_obj_v)):
                    black = True
            if ((black == False) & (checkifpallet(i,3) == False)):
                message = (str(cell_obj_v) +
                        ' ' + str(((sheet_obj.cell(row = i, column = 3)).value)) +
                        ' ' + str(((sheet_obj.cell(row = i, column = 4)).value)))
                black = False
                save_file.write((message + '\n'))
                print(message)
    while(True):
        menu = input('Please enter 0 to go to main menu and save results')
        if (menu == '0'):
            save_file.close()
            break
        else:
            print('Wrong action, please try again')

def findlocation():
    wb_obj.active = 2
    sheet_obj = wb_obj.active
    pallet = True
    pallet_v = input('Skip the pallets? ')
    if ((str(pallet_v) == 'yes') |  (str(pallet_v) == 'Yes') |  (str(pallet_v) == 'YES')):
        pallet = False
    else:
        pallet = True
    while(True):
        print('Please enter 0 to go to main menu')
        SKU = input('Please enter product SKU ')
        print("Item name   Location   On hand   Avaible")
        for i in range(2, len(list(sheet_obj.rows))):
            cell_obj = sheet_obj.cell(row = i, column = 1)
            cell_obj_v = cell_obj.value
            hand = ((sheet_obj.cell(row = i, column = col_on_hand)).value)
            if ((int(hand) >= 0) & (int(hand) <= 9)):
                space = '         '
            elif ((int(hand) >= 10) & (int(hand) <= 99)):
                space = '        '
            else:
                space = '       '
            if(pallet == True):
                if ((str(SKU) == str(cell_obj_v)) |
                    (str(SKU) == str(((sheet_obj.cell(row = i, column = col_UPC)).value))) |
                    (str(SKU) == str(((sheet_obj.cell(row = i, column = col_bin)).value)))):
                    print(str(cell_obj_v) +
                        '     ' + str(((sheet_obj.cell(row = i, column = col_bin)).value)) +
                        '    ' + str(((sheet_obj.cell(row = i, column = col_on_hand)).value)) +
                        space + str(((sheet_obj.cell(row = i, column = col_avaible)).value)))
            else:
                if(checkifpallet(i,3) == False):
                    if ((str(SKU) == str(cell_obj_v)) |
                        (str(SKU) == str(((sheet_obj.cell(row = i, column = col_UPC)).value))) |
                        (str(SKU) == str(((sheet_obj.cell(row = i, column = col_bin)).value)))):
                        print(str(cell_obj_v) +
                            '     ' + str(((sheet_obj.cell(row = i, column = col_bin)).value)) +
                            '    ' + str(((sheet_obj.cell(row = i, column = col_on_hand)).value)) +
                            space + str(((sheet_obj.cell(row = i, column = col_avaible)).value)))
        if (SKU == '0'):
            break

def diff_hand_avaible():
    wb_obj.active = 2
    sheet_obj = wb_obj.active
    for i in range(1, len(list(sheet_obj.rows))):
        hand = str(((sheet_obj.cell(row = i, column = col_on_hand)).value))
        avaible = str(((sheet_obj.cell(row = i, column = col_avaible)).value))
        if (hand != avaible):
            print(str(sheet_obj.cell(row = i, column = col_name).value) +
                ' ' + str(((sheet_obj.cell(row = i, column = col_bin)).value)) +
                ' ' + str(hand) +
                ' ' + str(avaible))
    while(True):
        menu = input('Please enter 0 to go to main menu and save results')
        if (menu == '0'):
            break
        else:
            print('Wrong action, please try again')
def sold():
    wb_obj.active = 3
    sheet_obj = wb_obj.active
    SKU = input('Please enter product SKU ')
    for i in range(2, len(list(sheet_obj.rows))):
        cell_obj = sheet_obj.cell(row = i, column = 1)
        cell_obj_v = cell_obj.value
        if ((str(SKU) == str(cell_obj_v)) |
            (str(SKU) == str(((sheet_obj.cell(row = i, column = 2)).value)))):
            print('Item ' + str(cell_obj_v) + ' was sold ' + )
                    str(((sheet_obj.cell(row = i, column = 2)).value)) + 'times last 3 months')
    while(True):
        menu = input('Please enter 0 to go to main menu and save results')
        if (menu == '0'):
            break
        else:
            print('Wrong action, please try again')

while(True):
    print('1 - Inventory')
    print('2 - Replenishment')
    print('3 - Double locations')
    print('4 - Test func')
    print('5 - Find location')
    print('6 - Diffrence "on hand" and "avaible" ')
    print('9 - Close application')
    print('0 - Main menu')
    menu = input('Please enter action:')
    if (menu == '9'):
        file1.close()
        break
    elif (menu == '1'):
        inventory()
    elif (menu == '2'):
        replenishment()
    elif (menu == '3'):
        doublelocations()
    elif (menu == '4'):
        testfunc()
    elif (menu == '5'):
        findlocation()
    elif (menu == '6'):
        diff_hand_avaible()
    elif (menu == '7'):
        sold()
    else:
        print('Wrong Action, please try again')
